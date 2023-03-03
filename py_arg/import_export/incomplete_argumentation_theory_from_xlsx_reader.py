from pathlib import Path
from typing import Union

import pandas as pd
from openpyxl import load_workbook

from py_arg.aspic_classes.argumentation_system import ArgumentationSystem
from py_arg.aspic_classes.defeasible_rule import DefeasibleRule
from py_arg.aspic_classes.literal import Literal
from py_arg.incomplete_aspic_classes.incomplete_argumentation_theory import IncompleteArgumentationTheory


class IncompleteArgumentationTheoryFromXLSXFileReader:
    def __init__(self):
        pass

    @staticmethod
    def read_from_xlsx_file(file_path: Union[Path, str]) -> IncompleteArgumentationTheory:
        wb_sheet_names = load_workbook(file_path, read_only=True).sheetnames

        _literal_df = pd.read_excel(file_path, sheet_name='Literals')
        _rule_df = pd.read_excel(file_path, sheet_name='Rules')
        _query_df = pd.read_excel(file_path, sheet_name='Queries', index_col='Query')

        if 'Contraries' in wb_sheet_names:
            _contrary_df = pd.read_excel(file_path, sheet_name='Contraries')
        else:
            _contrary_df = None
        if 'RulePreferences' in wb_sheet_names:
            _rule_preference_df = pd.read_excel(file_path, sheet_name='RulePreferences')
        else:
            _rule_preference_df = None

        try:
            language = dict()
            queryables = []
            contraries_and_contradictories = {}

            for index, row in _literal_df.iterrows():
                literal = Literal(row['Literal'])
                neg_literal = Literal('-' + row['Literal'])
                language[literal.s1] = literal
                language[neg_literal.s1] = neg_literal
                if row['Observable'] == 'y':
                    queryables.append(literal)
                    queryables.append(neg_literal)
                contraries_and_contradictories[literal.s1] = {neg_literal}
                contraries_and_contradictories[neg_literal.s1] = {literal}

            # Add additional contraries
            if _contrary_df is not None:
                for index, row in _contrary_df.iterrows():
                    literal_str = row['Literal'].replace('~', '-')
                    contraries_str = [cont.strip().replace('~', '-') for cont in row['Contraries'].split(',')]
                    for contrary_str in contraries_str:
                        contraries_and_contradictories[literal_str].add(language[contrary_str])

            # Add rules
            rules = []
            for index, row in _rule_df.iterrows():
                ants_str = [ant.strip().replace('~', '-') for ant in row['Antecedents'].split(',')]
                cons_str = row['Consequent'].strip().replace('~', '-')
                if any([ant not in language for ant in ants_str]) or cons_str not in language:
                    raise KeyError(
                        'Not each literal in ' + str(ants_str) + ' or ' + cons_str + ' was in logical language.')
                ants = {language[literal_str] for literal_str in ants_str}
                cons = language[cons_str]

                if 'ID' in _rule_df.keys():
                    rule_id = row['ID']
                else:
                    rule_id = index

                rules.append(DefeasibleRule(rule_id, ants, cons))

            argumentation_system = ArgumentationSystem(language=language,
                                                       contraries_and_contradictories=contraries_and_contradictories,
                                                       strict_rules=[], defeasible_rules=rules,
                                                       add_defeasible_rule_literals=False)
            return IncompleteArgumentationTheory(argumentation_system=argumentation_system,
                                                 queryables=queryables,
                                                 knowledge_base_axioms=[],
                                                 knowledge_base_ordinary_premises=[])
        except Exception as exception:
            raise ImportError('Could not load Argumentation System.' + str(type(exception)))
